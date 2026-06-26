#!/usr/bin/env python3
"""build-link-report.py — turn the hyperlink candidate log into a review workbook.

Reads data/links/link-candidates.csv (written by apply-hyperlinks.py — one row per
matched occurrence, applied or not) and writes data/links/hyperlink-opportunities.xlsx:

    Summary        counts by family / applied-state / confidence band
    Applied        every link applied to the pages (confidence >= 0.75)
    Below bar      every candidate NOT applied (< 0.75) — to mine for more later
    All candidates  the full log

Autofilter + frozen header + sensible widths on each sheet.

Usage:  python3 scripts/build-link-report.py
"""
from __future__ import annotations
import csv
from collections import Counter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CSV = ROOT / "data" / "links" / "link-candidates.csv"
OUT = ROOT / "data" / "links" / "hyperlink-opportunities.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="22303C")
HEAD_FONT = Font(bold=True, color="FFFFFF")
APPLIED_FILL = PatternFill("solid", fgColor="E7F3E7")
BELOW_FILL = PatternFill("solid", fgColor="FBEFE6")
WIDTHS = [34, 6, 22, 18, 16, 34, 40, 10, 9, 26, 60]


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"
    for i, w in enumerate(WIDTHS[:ncols], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_sheet(wb, title, header, rows, fill=None):
    ws = wb.create_sheet(title)
    ws.append(header)
    for r in rows:
        ws.append(r)
    # tint only the (cheap) applied/confidence column instead of every cell
    if fill:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=9).fill = fill   # "applied" column
    style_header(ws, len(header))
    return ws


def main():
    rows = list(csv.DictReader(CSV.open(encoding="utf-8")))
    header = ["source_file", "line", "matched_text", "surface_form", "target_family",
              "target_file", "target_anchor", "confidence", "applied", "reason", "context"]

    def tup(r):
        return [r["source_file"], int(r["line"]), r["matched_text"], r["surface_form"],
                r["target_family"], r["target_file"], r["target_anchor"],
                float(r["confidence"]), r["applied"], r["reason"], r["context"]]

    applied = [tup(r) for r in rows if r["applied"] != "no"]
    below = [tup(r) for r in rows if r["applied"] == "no"]
    allr = [tup(r) for r in rows]
    applied.sort(key=lambda x: (x[4], -x[7], x[0]))
    below.sort(key=lambda x: (-x[7], x[4], x[0]))
    allr.sort(key=lambda x: (x[0], x[1]))

    wb = Workbook()
    wb.remove(wb.active)

    # summary
    ws = wb.create_sheet("Summary")
    ws.append(["Hyperlink opportunities — summary"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Total candidates", len(rows)])
    ws.append(["Applied (>= 0.75)", len(applied)])
    ws.append(["Below bar (< 0.75)", len(below)])
    ws.append([])
    ws.append(["By family", "applied", "below bar"])
    fam_a = Counter(r["target_family"] for r in rows if r["applied"] != "no")
    fam_b = Counter(r["target_family"] for r in rows if r["applied"] == "no")
    for fam in sorted(set(fam_a) | set(fam_b)):
        ws.append([fam, fam_a.get(fam, 0), fam_b.get(fam, 0)])
    ws.append([])
    ws.append(["By confidence band", "count"])
    bands = Counter()
    for r in rows:
        c = float(r["confidence"])
        b = ">=0.90" if c >= 0.9 else ">=0.80" if c >= 0.8 else ">=0.75" if c >= 0.75 \
            else ">=0.60" if c >= 0.6 else "<0.60"
        bands[b] += 1
    for b in [">=0.90", ">=0.80", ">=0.75", ">=0.60", "<0.60"]:
        ws.append([b, bands.get(b, 0)])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    for row in ws.iter_rows(min_row=7, max_row=7):
        for cell in row:
            cell.font = Font(bold=True)
    ws["A10"].font = Font(bold=True)

    write_sheet(wb, "Applied", header, applied, APPLIED_FILL)
    write_sheet(wb, "Below bar", header, below, BELOW_FILL)
    write_sheet(wb, "All candidates", header, allr)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(ROOT)}")
    print(f"  {len(rows)} candidates | {len(applied)} applied | {len(below)} below bar")


if __name__ == "__main__":
    main()
