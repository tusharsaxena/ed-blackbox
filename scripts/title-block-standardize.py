#!/usr/bin/env python3
"""
title-block-standardize.py — Build the title-block standardization workbook.

Reads:
  scripts/out/title-blocks.json     (current state, from extract-title-blocks.py)
  scripts/out/git-dates.json        (git last-modified fallback dates)
  scripts/out/proposed-overrides.json  (OPTIONAL — per-page judgment proposals
                                         emitted by the Phase-1 workflow, keyed by file)

Writes:
  scripts/out/title-blocks.xlsx     (one row per page: CURRENT vs PROPOSED columns,
                                      a Δ change-flag, plus a Rules sheet)

The PROPOSED columns are a deterministic first pass from the per-group canonical
schema below, with workflow judgment proposals merged on top where provided.
Run AFTER extract-title-blocks.py.  Paths resolve relative to the repo root.
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "scripts" / "out"

# ---- The standard: semantic title-block columns -------------------------------
# Order matters — this is the column order in the sheet.
ELEMENTS = [
    ("page_name",    "Page name (subject)"),
    ("title_suffix", "<title> group suffix"),
    ("title_full",   "<title> full"),
    ("kicker_a",     "Kicker — label A"),
    ("kicker_b",     "Kicker — label B"),
    ("h1_lead",      "H1 lead text"),
    ("h1_accent",    "H1 accent <span>"),
    ("h1_role",      "H1 .role tag"),
    ("subtitle",     "Subtitle (drop)"),
    ("meta_label",   "Meta — label"),
    ("meta_updated", "Meta — Updated"),
]

# ---- Per-group canonical rules (deterministic baseline) -----------------------
# Fields set to None need per-page judgment (filled from overrides / left as current).
CANON = {
    "index":       dict(title_suffix="",              kicker_a="Elite Dangerous", meta_label=None),
    "engineering": dict(title_suffix="Engineering",   kicker_a="Engineering",     meta_label="Series Engineering"),
    "systems":     dict(title_suffix="Systems",       kicker_a="Systems",         meta_label="Series Systems"),
    "ships":       dict(title_suffix="Ships",         kicker_a="Guide",           meta_label="Series Ships"),
    "by-role":     dict(title_suffix="Best Ships",    kicker_a="Role Dossier",    meta_label="Best Ships by Role"),
    "dossier":     dict(title_suffix=None,            kicker_a="Ship &times; Role Dossier", meta_label="Ship &times; Role Dossier"),
    "activities":  dict(title_suffix="Role &amp; Activities", kicker_a="Role &amp; Activities", meta_label="Role &amp; Activities"),
}


def load(name, default):
    p = OUT / name
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else default


def proposed_for(cur, overrides, git_dates):
    """Deterministic proposed row, then merged with any workflow override."""
    grp = cur["group"]
    canon = CANON.get(grp, {})
    f = cur["file"]

    # page_name: the subject — prefer the existing <title> name part (clean, drives
    # title_full), else breadcrumb, else accent span
    tname, _tsuf = split_title(cur.get("title_tag", ""))
    page_name = tname or cur.get("breadcrumb_current") or cur.get("h1_accent") or ""

    # baseline from canon + current
    p = {
        "page_name":    page_name,
        "title_suffix": canon.get("title_suffix"),
        "kicker_a":     canon.get("kicker_a"),
        "kicker_b":     cur["kicker_parts"][1] if len(cur["kicker_parts"]) > 1 else (cur["kicker_parts"][0] if cur["kicker_parts"] else ""),
        "h1_lead":      cur.get("h1_lead") or "",  # preserve; judgment agents may trim
        "h1_accent":    cur.get("h1_accent") or page_name,
        "h1_role":      cur.get("h1_role") or "",
        "subtitle":     "",                       # standard: masthead carries no subtitle
        "meta_label":   canon.get("meta_label"),
        "meta_updated": cur.get("updated_date") or git_dates.get(f, ""),
    }

    # per-group fills where canon left None
    if grp == "dossier":
        p["title_suffix"] = cur.get("h1_role") or ""        # suffix = role
    if grp == "by-role":
        p["kicker_b"] = cur.get("h1_role")  # role lives in kicker_b already
        p["kicker_b"] = cur["kicker_parts"][1] if len(cur["kicker_parts"]) > 1 else ""
        p["h1_role"] = "Ship Comparison"
    if grp == "index":
        p["meta_label"] = cur.get("meta_label") or ""
        p["kicker_b"] = "Guides"
        p["h1_accent"] = cur.get("h1_accent") or "Black Box"

    # merge workflow judgment override: any key PRESENT in the override wins, even
    # if explicitly blank (judged rows fully specify intent, e.g. clearing h1_lead).
    ov = overrides.get(f, {})
    for k, v in ov.items():
        if k in p and v is not None:
            p[k] = v

    # title_full assembled last
    suffix = p["title_suffix"]
    if grp == "index":
        p["title_full"] = "Elite:Dangerous Black Box"
    elif suffix:
        p["title_full"] = f'{p["page_name"]} &middot; {suffix} | E:D Black Box'
    else:
        p["title_full"] = f'{p["page_name"]} | E:D Black Box'

    # carry through any flags/notes from the override
    p["_flags"] = ov.get("flags", "")
    return p


def split_title(title_tag):
    """Parse '<name> &middot; <suffix> | E:D Black Box' -> (name, suffix)."""
    t = title_tag
    if "|" in t:
        t = t.split("|", 1)[0].strip()
    name, suffix = t, ""
    for seph in ("&middot;", "·"):
        if seph in t:
            name, suffix = [x.strip() for x in t.split(seph, 1)]
            break
    return name, suffix


def current_view(cur):
    """Map the raw extracted row onto the same semantic columns for side-by-side."""
    kp = cur["kicker_parts"]
    name, suffix = split_title(cur.get("title_tag", ""))
    return {
        "page_name":    name,
        "title_suffix": suffix,
        "title_full":   cur.get("title_tag", ""),
        "kicker_a":     kp[0] if kp else "",
        "kicker_b":     kp[1] if len(kp) > 1 else "",
        "h1_lead":      cur.get("h1_lead", ""),
        "h1_accent":    cur.get("h1_accent", ""),
        "h1_role":      cur.get("h1_role", ""),
        "subtitle":     cur.get("subtitle", ""),
        "meta_label":   cur.get("meta_label", ""),
        "meta_updated": cur.get("updated_date", ""),
    }


def main():
    rows = load("title-blocks.json", [])
    git_dates = load("git-dates.json", {})
    overrides = load("proposed-overrides.json", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Title Blocks"

    # ---- styling
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    cur_fill = PatternFill("solid", fgColor="EEF2F7")
    prop_fill = PatternFill("solid", fgColor="E8F5E9")
    chg_fill = PatternFill("solid", fgColor="FFF3CD")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    grp_font = Font(bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    # ---- header rows (two-tier: CURRENT | PROPOSED)
    cols = ["file", "group", "Δ"]
    cur_cols = [f"cur:{k}" for k, _ in ELEMENTS]
    prop_cols = [f"new:{k}" for k, _ in ELEMENTS]
    cols += cur_cols + prop_cols + ["flags / notes"]

    ws.append(cols)
    for ci, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = wrap
        c.border = border
        if name.startswith("cur:"):
            c.value = name[4:]
        elif name.startswith("new:"):
            c.value = name[4:]

    n_changed = 0
    for cur in rows:
        view = current_view(cur)
        prop = proposed_for(cur, overrides, git_dates)
        # change detection across the semantic elements
        changed_keys = [k for k, _ in ELEMENTS
                        if (view.get(k, "") or "").strip() != (str(prop.get(k, "") or "")).strip()]
        delta = "●" if changed_keys else ""
        if changed_keys:
            n_changed += 1

        line = [cur["file"], cur["group"], delta]
        line += [view.get(k, "") for k, _ in ELEMENTS]
        line += [prop.get(k, "") for k, _ in ELEMENTS]
        line += [prop.get("_flags", "")]
        ws.append(line)
        r = ws.max_row
        base = 4  # first cur col index
        for i, (k, _) in enumerate(ELEMENTS):
            ws.cell(row=r, column=base + i).fill = cur_fill
            pc = ws.cell(row=r, column=base + len(ELEMENTS) + i)
            pc.fill = chg_fill if k in changed_keys else prop_fill
        for ci in range(1, len(cols) + 1):
            ws.cell(row=r, column=ci).alignment = wrap
            ws.cell(row=r, column=ci).border = border

    # widths
    widths = [34, 11, 4] + [22] * len(ELEMENTS) + [22] * len(ELEMENTS) + [40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # ---- Rules sheet
    rs = wb.create_sheet("Rules")
    rules = [
        ["TITLE-BLOCK STANDARD — first pass (edit the `new:` columns to revise)", ""],
        ["", ""],
        ["Element", "Standard"],
        ["<title>", "<Page name> &middot; <group suffix> | E:D Black Box   (separator: &middot;)"],
        ["kicker", "Two segments joined by <span class=\"sep\">//</span>: <label A> // <label B>"],
        ["h1.title", "<span>accent word</span> + <span class=\"role\">descriptor</span>; drop leading 'The'"],
        ["subtitle", "DROP — the masthead carries no subtitle (CLAUDE.md). Content moves to the briefing."],
        ["masthead-meta", "Exactly 2 spans: <type label> + Updated <b>YYYY-MM-DD</b>. Extra bespoke spans dropped (see flags)."],
        ["", ""],
        ["Group", "suffix / kicker-A / meta-label  (fixed); judgment = kicker-B + .role"],
        ["dossier", "<role> / Ship × Role Dossier / Ship × Role Dossier ; kicker-B = manufacturer"],
        ["by-role", "Best Ships / Role Dossier / Best Ships by Role ; .role = Ship Comparison"],
        ["activities", "Role & Activities / Role & Activities / Role & Activities"],
        ["engineering", "Engineering / Engineering / Series Engineering ; kicker-B + .role per page"],
        ["systems", "Systems / Systems / Series Systems ; kicker-B + .role per page"],
        ["ships", "Ships / Guide / Series Ships"],
        ["index", "(special) Elite:Dangerous Black Box"],
        ["", ""],
        ["Legend", "Δ ● = proposed differs from current. Yellow new: cell = changed. Green = unchanged."],
        ["Dates", "7 systems pages had no Updated date; proposed value is the git last-commit date — VERIFY."],
    ]
    for row in rules:
        rs.append(row)
    rs.column_dimensions["A"].width = 18
    rs.column_dimensions["B"].width = 90
    rs.cell(row=1, column=1).font = Font(bold=True, size=12)
    for rr in (3, 10):
        rs.cell(row=rr, column=1).font = grp_font
        rs.cell(row=rr, column=2).font = grp_font

    OUT.mkdir(parents=True, exist_ok=True)
    wb.save(OUT / "title-blocks.xlsx")
    print(f"Wrote scripts/out/title-blocks.xlsx — {len(rows)} pages, {n_changed} with proposed changes")
    print(f"  overrides merged for {len(overrides)} pages")


if __name__ == "__main__":
    main()
